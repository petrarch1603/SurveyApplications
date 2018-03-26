# import xml.sax, xml.sax.handler
# from xml.dom import minidom

import openpyxl
import xml.etree.ElementTree as et
from zipfile import ZipFile

# This is a script for taking the status of rows on a tracking spreadsheet and updating a KML based on
# the fields in that tracking spreadsheet.

# Initialize Variables
mykmz = 'source_data/map.kmz'
myxlsx = 'source_data/tracking.xlsx'

# TODO Receive emails on Raspberry PI with attachments and process them with this program

# Read the Spreadsheet
wb = openpyxl.load_workbook(myxlsx)


# Create a dictionary with all the relevant spreadsheet data
sheet = wb.sheetnames[0]
sheet = wb[sheet]

pointdata = {}
for row in range(2, sheet.max_row + 1):
    pointno = sheet['A' + str(row)].value
    dict = {
        'typeset': sheet['H' + str(row)].value,
        'setby': sheet['I' + str(row)].value,
        'dateset': sheet['J' + str(row)].value,
        'color': sheet['H' + str(row)].fill.start_color.index
    }
    pointdata.setdefault(pointno, {})
    pointdata[pointno] = dict

# Unzip the KMZ and extract doc.kml


kmz = ZipFile(mykmz, 'r')
kml = kmz.open('doc.kml', 'r')

urltag = "{http://www.opengis.net/kml/2.2}"
tree = et.parse(kml)
root = tree.getroot()
d = tree.findall(".//" + urltag + "Placemark")

for i in d:
    for subelem in i:
        if subelem.tag == (urltag + "name"):
            for k, v in pointdata.items():
                if subelem.text == str(k):
                    i[2].text = ('Set on ' + str(v['dateset']))
tree = et.ElementTree(root)
tree.write("newkml2.kml", xml_declaration=True)




# TODO Use Filenames after UpdateKML.py as arguments

# TODO Import and parse excel spreadsheets

# TODO Import and process KMLs

# TODO Return the new KML with date as part of filename
